#!/usr/bin/env python
# coding: utf-8

# In[ ]:


versionString="24"                                   # version string of this application


# ## NanoDiP all-in-one Jupyter Notebook
# *J. Hench, S. Frank, and C. Hultschig, Neuropathology, IfP Basel, 2021*
# 
# This software is provided free of charge and warranty; by using it you agree to do this on your own risk. The authors shall not be held liable for any damage caused by this software. We have assembled this and tested it to the best of our knowledge.
# 
# The purpose of NanoDiP (Nanopore Digital Pathology) is to compare low-coverage Nanopore sequencing data from natively extracted DNA sequencing runs against a flexibly adaptable collection of 450K/850K Illumina Infinium Methylation array data. These data have to be preprocessed into binary beta value files; this operation is performed in R (uses minfi to read raw array data) and outputs bindary float files (one per dataset). These beta values files (e.g., 204949770141_R03C01_betas_filtered.bin) are named according to the array ID (Sentrix ID) followed by the suffix. A collection of betas_filtered.bin files can be provided in a static manner and XLSX (Microsoft Excel) tables can be used to select a subset thereof alongside a user-defined annotation. The corresponding datasets will be loaded into memory and then serve as the reference cohort to which the Nanopore data are compared by dimension reduction (UMAP). This comparison is optimized for speed and low resource consumption so that it can run on the computer that operates the sequencer. The sequencing run is initiated through the MinKNOW API by this application. Basecalling and methylation calling occur as background tasks outside this Jupyter Notebook. User interaction occurs through a web interface based on CherryPy which has been tested on Chromium web browser. It is advisable to run it locally, there are no measures to secure the generated website.
# 
# In order to use this application properly please make sure to be somewhat familiar with Jupyter Notebook. To run the software, press the button called *restart the kernel, re-run the whole notebook (with dialog)* and confirm execution. Then, in Chromium Browser, navigate to http://localhost:8080/ and preferably bookmark this location for convenience. In case of errors, you may just again click the same button *restart the kernel, re-run the whole notebook (with dialog)*.
# ___
# ### Technical Details
# * Tested with Python 3.7.5; 3.8.8 fails to load minknow_api in jupyter notebook.
# * Verified to run on Ubuntu 18.04/Jetpack on ARMv8 and x86_64 CPUs; not tested on Windows and Mac OS. The latter two platforms are unsupported, we do not intend to support them.
# * **CAUTION**: Requires a *patched* version of minknow api, file `[VENV]/lib/python3.7/site-packages/minknow_api/tools/protocols.py`. Without the patch, the generated fast5 sequencing data will be unreadable with f5c or nanopolish (wrong compression algorithm, which is the default in the MinKNOW backend).
# 

# In[ ]:


# verify running Python version (should be 3.7.5) and adjust jupyter notebook
import IPython
import os
from IPython.core.display import display, HTML      # set display witdth to 100%
display(HTML("<style>.container { width:100% !important; }</style>"))
os.system('python --version')


# ## Multithreading Options
# Depending on the number of parallel threads/cores of the underlying hardware, threading options for multithreaded modules need to be set as environment-specific parameters. One way to do so is through the *os* module.

# In[ ]:


# execution-wide multithreading options, set according to your hardware. Jetson AGX: suggest "2"
# needs to be set before importing other modules that query these parameters
import os
os.environ["NUMBA_NUM_THREADS"] = "2" # export NUMBA_NUM_THREADS=2
os.environ["OPENBLAS_NUM_THREADS"] = "2"
os.environ["MKL_NUM_THREADS"] = "2"


# ## Modules
# This section imports the required modules that should have been installed via pip. Other package managers have not been tested. To install packages, use the setup script provided with this software or, alternatively, install them one by one, ideally in a virtual python environment. Note that the MinKNOW API requires manual patching after installation with pip.

# In[ ]:


# Python modules to load
import argparse
import cherrypy
import datetime
import fnmatch
import logging
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.pyplot import figure
from minknow_api.manager import Manager
import minknow_api.statistics_pb2
import minknow_api.device_pb2
from minknow_api.tools import protocols
from numba import jit
import numpy
import openpyxl
import os
from os import listdir
import pandas
import plotly.express as px
import plotly.graph_objects as go
import psutil
import pysam
import random
import shutil
import string
import socket
import subprocess
import sys
import time # for development purposes in jupyter notebook (progress bars)
import timeit # benchmarking
from tqdm.notebook import tqdm, trange # for development purposes in jupyter notebook (progress bars)
#import umap # installed via pip install umap-learn; import moved to UMAP function
import webbrowser
from xhtml2pdf import pisa


# ## Configuration
# Below are system-specific parameters that may or may not require adaptation. Many variable names are be self-explanatory. The key difference between Nanopore setups are between devices provided by ONT (MinIT incl. running the MinIT distribution on a NVIDIA Jetson developer kit such as the AGX Xavier, GridION) and the typical Ubuntu-based MinKNOW version on x86_64 computers. The raw data are written into a `/data` directory on ONT-based devices while they are found in `/var/lib/minknow/data` on x86_64 installations. Make sure to adapt your `minknowDataDir` accordingly. There are furthermore permission issues and special folders / files in the MinKNOW data directory. These files / folders should be excluded from analysis through `fileHideList` so that only real run folders will be parsed. Finally, the `nanodipOutputDir` is the place in which the background methylation and alignment process will place its results by replicating the directory hierarchy of the MinKNOW data location. It will not duplicate the data, and these data will be much smaller than raw run data. They can be placed anywhere in the file tree, but also inside the MinKNOW data path within a sub-folder. If the latter is the case, make sure to apply appropriate read/write permissions. Final reports and figures generated by NanoDiP are written into `nanodipReportDir`.

# In[ ]:


# configuration parameters, modify here, no need for a configuration file
minknowDataDir="/data"                              # where MinKNOW places its data
fileHideList=["pings",                              # list of files and folders to be exluded from parsing
              "reads","queued_reads","core-dump-db","lost+found",
              "intermediate",
              "minimap_data","nanodip_tmp","nanodip_output",
              "nanodip_reports",
              "non-ont","raw_for_playback","user_scripts",
              "playback_raw_runs",".Trash-1000"]
nanodipOutputDir="/data/nanodip_output"             # location to write intermediate analysis data, i.e. methylation and alignment files
nanodipReportDir="/data/nanodip_reports"            # location to write reports and figures
readsPerFile="400"                                  # number of reads per file. 400 works well on the Jetson AGX. Higher numbers increase batch size and RAM usage, lower numbers use more I/O resouces due to more frequent reloading of alignment reference
wantedBases=150000000                               # number of basecalled bases until run termination occurs
resultEndings=["_UMAP_top.html","_UMAP_all.html",   # list of file name sections that identify past runs
               "_NanoDiP_report.pdf","_CNVplot.png",
               "_NanoDiP_ranking.pdf"]
analysisExclude=["_TestRun_"]                       # string patterns in sample names that exclude data from downstream analysis, e.g., test runs
thisFaviconPath="/applications/nanodip/favicon.ico" # the web browser favicon file for this application
epidipLogoPath="/applications/nanodip/EpiDiP_Logo_01.png" # logo bitmap for PDF reports
imgPath="/applications/nanodip"                     # the location where image files for the web application are stored
cpgScriptPath="/applications/nanodip/calculate_overlap_CpGs_04.sh" # script that loops over all fast5 folders, screens for predominant barcode and launched minimap2/f5c for each fast5 file 
thisHost="localhost"                                # name of the host computer, typically "localhost"
cherrypyHost="localhost"                            # name of the host, typically "localhost" as well
cherrypyPort=8080                                   # port on which the NanoDiP UI will be served
cherrypyThreads=100                                 # number of concurrent threads allowed to CherryPy, decrease in case of performance problems, default = 100
debugLogging=False                                  # CherryPy debug logging incl. access logs (True for testing only)
binDir="/applications/reference_data/betaEPIC450Kmix_bin" # location of the preprocessed beta value data
binIndex=binDir+"/index.csv"                        # the index file of the beta value binary files is stored in a CSV files and is generated by the same R script that creates the beta value float binary files
referenceDir="/applications/reference_data/reference_annotations" # location of the XLSX files that contain annotations, i.e. reference file collection definitions
methylCutOff=0.35                                   # cut-off for "unmethylated vs methylated" for Illumina array data; also applicable to other methylation data types
topMatch=100                                        # number of reference cases to be shown in subplot including copy number profile links (not advisable >200, plotly will become really slow)
cnvLinkPrefix="http://s1665.rootserver.io/umapplot01/" # URL prefix to load PDF with CNV plot for a given Sentrix ID
cnvLinkSuffix="_CNV_IFPBasel_annotations.pdf"       # URL prefix to load PDF with CNV plot 
chrLengthsFile="/applications/reference_data/hg19_cnv/ChrLengths_hg19.tsv"          # contains three columns, A:chrom. strings, B: chrom. lengths, C: offsets
centromereLocationsBed="/applications/reference_data/hg19_cnv/hg19.centromere.bed"  # contains the centromere positions for each chromosome
plotlyRenderMode="webgl"                            # default="webgl", alternative "svg" without proper webgl support (e.g., firefox, use "svg"; slower, but does not require GPU)
barcodeNames=["barcode01","barcode02","barcode03",  # barcode strings, currently kit SQK-RBK004
              "barcode04","barcode05","barcode06",
              "barcode07","barcode08","barcode09",
              "barcode10","barcode11","barcode12"]
refgenomefa="/applications/reference_data/minimap_data/hg19.fa" # human reference genome
refgenomemmi="/applications/reference_data/minimap_data/hg19_20201203.mmi" # human reference genome minimap2 mmi
ilmncgmapfile="/applications/reference_data/minimap_data/hg19_HumanMethylation450_15017482_v1-2_cgmap.tsv" # Illumina probe names of the 450K array
f5cBin="/applications/f5c/f5c"                      # f5c binary location (absolute path) v6
minimap2Bin="/applications/nanopolish/minimap2/minimap2" # minimap2 binary location (absolute path)
samtoolsBin="/applications/samtools/samtools"       # samtools binary location (absolute path)
rscriptBin="/applications/R-4.0.3/bin/Rscript"      # Rscript binary location (absolute path)
readCpGscript="/applications/nanodip/readCpGs_mod02.R" # R script that reads CpGs into simplified text file (absolute path)
verbosity=0                                         # 0=low log verbosity, 1=high log verbosity (with timestamps, for benchmarking and debugging)


# # No user editable code below
# Do not modify the cells below unless you would like to patch errors or create something new.
# ## Sections
# 1. Generic Functions
# 2. MinKNOW API Functions
# 3. CNV Plotter
# 4. UMAP Methylation Plotter
# 5. User Interface Functions
# 6. Report Generator
# 7. CherryPy Web UI

# ### 1. Generic Functions

# In[ ]:


def logpr(v,logstring): # logging funcion that reads verbosity parameter
    if v==1:
        print(str(datetime.datetime.now())+": "+str(logstring))


# In[ ]:


def restartNanoDiP():
    cherrypy.engine.restart()


# In[ ]:


def getRuns(): # lists run folders from MinKNOW data directory in reverse order based on modif. date
    runFolders=[]
    #runFolderDates=[]
    for r in listdir(minknowDataDir):
        if r not in fileHideList:
            f=minknowDataDir+"/"+r
            if os.path.isdir(f):
                runFolders.append([r,float(os.path.getmtime(f))])
    runFolders.sort(key=lambda row: (row[1], row[0]), reverse=True) # sort based on modif. date
    runFolders=[j.pop(0) for j in runFolders] # remove date column after sorting
    return(runFolders)


# In[ ]:


def getPredominantBarcode(sampleName):
    fast5List = [os.path.join(dp, f) for dp, dn, filenames in os.walk(minknowDataDir+"/"+sampleName) for f in filenames if os.path.splitext(f)[1] == '.fast5']
    barcodeHits=[]
    for b in range(len(barcodeNames)):
        c=0
        for f in fast5List:
            if barcodeNames[b] in f:
                c+=1
        barcodeHits.append(c)
    maxbarcode=max(barcodeHits)
    if maxbarcode>1:
        predominantBarcode=barcodeNames[barcodeHits.index(maxbarcode)]
    else:
        predominantBarcode="undetermined"
    return predominantBarcode


# In[ ]:


def datetimestringnow(): # get current date and time as string to create timestamps
    now = datetime.datetime.now()
    return str(now.year).zfill(4)+str(now.month).zfill(2)+str(now.day).zfill(2)+"_"+str(now.hour).zfill(2)+str(now.minute).zfill(2)+str(now.second).zfill(2)  


# In[ ]:


def convert_html_to_pdf(source_html, output_filename): # generate reports
    result_file = open(output_filename, "w+b")         # open output file for writing (truncated binary)
    pisa_status = pisa.CreatePDF(                      # convert HTML to PDF
            source_html,                               # the HTML to convert
            dest=result_file)                          # file handle to recieve result
    result_file.close()                                # close output file
    return pisa_status.err                            # return True on success and False on errors


# In[ ]:


def getReferenceAnnotations(): # list all reference annotation files (MS Excel XLSX format)
    referenceAnnotations=[]
    for r in listdir(referenceDir):
        if r.endswith('.xlsx'):
            referenceAnnotations.append(r)    
    return referenceAnnotations


# In[ ]:


def writeReferenceDefinition(sampleId,referenceFile): # write the filename of the UMAP reference for the 
    with open(nanodipReportDir+'/'+sampleId+'_selected_reference.txt', 'w') as f: # current run into a text file
        f.write(referenceFile)


# In[ ]:


def readReferenceDefinition(sampleId): # read the filename of the UMAP reference for the current sample
    try:
        with open(nanodipReportDir+'/'+sampleId+'_selected_reference.txt', 'r') as f:
            referenceFile=f.read()
    except:
        referenceFile=""
    return referenceFile


# In[ ]:


def writeRunTmpFile(sampleId,deviceId):
    with open(nanodipReportDir+'/'+sampleId+'_'+deviceId+'_runinfo.tmp', 'a') as f: # current run into a text file
        try:
            runId=getActiveRun(deviceId)
        except:
            runId="none"
        ro=getThisRunOutput(deviceId,sampleId,runId)
        readCount=ro[0]
        bascalledBases=ro[1]
        overlapCpGs=getOverlapCpGs(sampleId)
        f.write(str(int(time.time()))+"\t"+
                str(readCount)+"\t"+
                str(bascalledBases)+"\t"+
                str(overlapCpGs)+"\n")


# In[ ]:


def readRunTmpFile(sampleId):
    print("readRunTmpFile not ready")


# In[ ]:


def getOverlapCpGs(sampleName):
    methoverlapPath=nanodipOutputDir+"/"+sampleName # collect matching CpGs from sample
    methoverlapTsvFiles=[] # find all *methoverlap.tsv files
    for root, dirnames, filenames in os.walk(methoverlapPath):
        for filename in fnmatch.filter(filenames, '*methoverlap.tsv'):
            methoverlapTsvFiles.append(os.path.join(root, filename))
    methoverlap=[]
    first=True
    for f in methoverlapTsvFiles:
        try: # some fast5 files do not contain any CpGs
            m=pandas.read_csv(f, delimiter='\t', header=None, index_col=0)
            if first:
                methoverlap=m
                first=False
            else:
                methoverlap=methoverlap.append(m)
        except:
            logpr(verbosity,"empty file encountered, skipping")
    return len(methoverlap)


# In[ ]:


def f5cOneFast5(sampleId,analyzeOne=True):
    analyzedCount=0
    thisRunDir=minknowDataDir+"/"+sampleId
    pattern = '*.fast5'
    fileList = []
    for dName, sdName, fList in os.walk(thisRunDir): # Walk through directory
        for fileName in fList:
            if fnmatch.fnmatch(fileName, pattern): # Match search string
                fileList.append(os.path.join(dName, fileName))
    calledList=[]
    completedCount=0
    maxBcCount=1 # at least 2 "passed" files (>1) need to be present
    targetBc="undetermined"
    for bc in barcodeNames:
        thisBc=0
        for f in fileList:
            if bc in f:
                if "_pass_" in f:
                    thisBc+=1
        if thisBc > maxBcCount:
            maxBcCount=thisBc
            targetBc=bc
    f5cAnalysisDir=nanodipOutputDir+"/"+sampleId
    if os.path.exists(f5cAnalysisDir)==False:
        os.mkdir(f5cAnalysisDir)
    thisBcFast5=[]
    thisBcFastq=[]
    for f in fileList:
        if targetBc in f:
            q=f.replace(".fast5","").replace("fast5_pass","fastq_pass")+".fastq"
            if os.path.exists(q): # check if accompanying fastq exists
                thisBcFast5.append(f)
                thisBcFastq.append(q)
                thisBcFileName=f.split("/")
                thisBcFileName=thisBcFileName[len(thisBcFileName)-1].replace(".fast5","") # get name prefix (to be the analysis subdir name later)
                thisAnalysisDir=f5cAnalysisDir+"/"+thisBcFileName
                if os.path.exists(thisAnalysisDir)==False:
                    os.mkdir(thisAnalysisDir)
                target5=thisAnalysisDir+"/"+thisBcFileName+".fast5"
                targetq=thisAnalysisDir+"/"+thisBcFileName+".fastq"
                if os.path.exists(target5)==False:
                    os.symlink(f,target5)             # fast5 symlink
                if os.path.exists(targetq)==False:
                    os.symlink(q,targetq)             #fastq symlink
                if os.path.exists(thisAnalysisDir+"/"+thisBcFileName+"-methoverlapcount.txt")==False:
                    if (analyzeOne==True and analyzedCount==0) or analyzeOne==False:
                        cmd=f5cBin+" index -t 1 --iop 100 -d "+thisAnalysisDir+" "+targetq
                        p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE) #index, call methylation and get methylation frequencies
                        p.wait()
                        cmd=minimap2Bin+" -a -x map-ont "+refgenomemmi+" "+targetq+" -t 4 | "+samtoolsBin+" sort -T tmp -o "+thisAnalysisDir+"/"+thisBcFileName+"-reads_sorted.bam"
                        p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE) # get sorted BAM (4 threads)
                        p.wait()
                        cmd=samtoolsBin+" index "+thisAnalysisDir+"/"+thisBcFileName+"-reads_sorted.bam"
                        p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE) # index BAM
                        p.wait()
                        cmd=f5cBin+" call-methylation -B2000000 -K400 -b "+thisAnalysisDir+"/"+thisBcFileName+"-reads_sorted.bam -g "+refgenomefa+" -r "+targetq+" > "+thisAnalysisDir+"/"+thisBcFileName+"-result.tsv"
                        p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE) # set B to 2 megabases (GPU) and 0.4 kreads
                        p.wait()
                        cmd=f5cBin+" meth-freq -c 2.5 -s -i "+thisAnalysisDir+"/"+thisBcFileName+"-result.tsv > "+thisAnalysisDir+"/"+thisBcFileName+"-freq.tsv"
                        p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
                        p.wait()
                        cmd=rscriptBin+" "+readCpGscript+" "+thisAnalysisDir+"/"+thisBcFileName+"-freq.tsv "+ilmncgmapfile+" "+thisAnalysisDir+"/"+thisBcFileName+"-methoverlap.tsv "+thisAnalysisDir+"/"+thisBcFileName+"-methoverlapcount.txt"
                        p = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
                        p.wait()
                        calledList.append(thisBcFileName)
                        analyzedCount+=1
                else:
                    completedCount+=1
    return "Target = "+targetBc+"<br>Methylation called for "+str(calledList)+". "+str(completedCount+analyzedCount)+"/"+str(len(thisBcFast5))  


# ### 2. MinKNOW API Functions
# Check https://github.com/nanoporetech/minknow_api for reference.
# 
# The following code requires a patched version of the MinKNOW API, install it from https://github.com/neuropathbasel/minknow_api.

# In[ ]:


def mkManager(): # Construct a manager using the host + port provided. This is used to connect to
    return Manager(host=thisHost, port=9501, use_tls=False) # the MinKNOW service trough the MK API.


# In[ ]:


def listMinionPositions(): # list MinION devices that are currenty connected to the system 
    manager = mkManager()
    positions = manager.flow_cell_positions() # Find a list of currently available sequencing positions.  
    return(positions)   # User could call {pos.connect()} here to connect to the running MinKNOW instance.


# In[ ]:


def listMinionExperiments(): # list all current and previous runs in the MinKNOW buffer, lost after MinKNOW restart
    manager=mkManager()
    htmlHost="<b>Host: "+thisHost+"</b><br><table border='1'><tr>"
    positions=manager.flow_cell_positions() # Find a list of currently available sequencing positions. 
    htmlPosition=[]
    for p in positions:
        htmlPosinfo="<b>-"+str(p)+"</b><br>"
        connection = p.connect()
        mountedFlowCellID=connection.device.get_flow_cell_info().flow_cell_id # return the flow cell info
        htmlPosinfo=htmlPosinfo+"--mounted flow cell ID: <b>" + mountedFlowCellID +"</b><br>"
        htmlPosinfo=htmlPosinfo+"---"+str(connection.acquisition.current_status())+"<br>" # READY, STARTING, sequencing/mux = PROCESSING, FINISHING; Pause = PROCESSING
        protocols = connection.protocol.list_protocol_runs()
        bufferedRunIds = protocols.run_ids
        for b in bufferedRunIds:
            htmlPosinfo=htmlPosinfo+"--run ID: " + b +"<br>"
            run_info = connection.protocol.get_run_info(run_id=b)
            htmlPosinfo=htmlPosinfo+"---with flow cell ID: " + run_info.flow_cell.flow_cell_id +"<br>"
        htmlPosition.append(htmlPosinfo)
    hierarchy = htmlHost
    for p in htmlPosition:
        hierarchy=hierarchy + "<td valign='top'><tt>"+p+"</tt></td>"
    hierarchy=hierarchy+"</table>"
    return(hierarchy)


# In[ ]:


def getFlowCellID(thisDeviceId): # determine flow cell ID (if any). Note that some CTCs have an empty ID string.
    mountedFlowCellID="no_flow_cell"
    manager=mkManager()
    positions=manager.flow_cell_positions() # Find a list of currently available sequencing positions.
    for p in positions:
        if thisDeviceId in str(p):
            connection = p.connect()
            mountedFlowCellID=connection.device.get_flow_cell_info().flow_cell_id # return the flow cell info
    return mountedFlowCellID


# In[ ]:


# This cell starts a run on Mk1b devices and perform several checks concerning the run protocol.

# modified from the MinKNOW API on https://github.com/nanoporetech/minknow_api (2021-06)
# created from the sample code at
# https://github.com/nanoporetech/minknow_api/blob/master/python/examples/start_protocol.py
# minknow_api.manager supplies "Manager" a wrapper around MinKNOW's Manager gRPC API with utilities
# for querying sequencing positions + offline basecalling tools.
# from minknow_api.manager import Manager

# We need `find_protocol` to search for the required protocol given a kit + product code.
# from minknow_api.tools import protocols
def parse_args():
    """Build and execute a command line argument for starting a protocol.

    Returns:
        Parsed arguments to be used when starting a protocol.
    """
    parser = argparse.ArgumentParser(
        description="""
        Run a sequencing protocol in a running MinKNOW instance.
        """
    )
    parser.add_argument(
        "--host",
        default="localhost",
        help="IP address of the machine running MinKNOW (defaults to localhost)",
    )
    parser.add_argument(
        "--port",
        help="Port to connect to on host (defaults to standard MinKNOW port based on tls setting)",
    )
    parser.add_argument(
        "--no-tls", help="Disable tls connection", default=False, action="store_true"
    )
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging")

    parser.add_argument("--sample-id", help="sample ID to set")
    parser.add_argument(
        "--experiment-group",
        "--group-id",
        help="experiment group (aka protocol group ID) to set",
    )
    parser.add_argument(
        "--position",
        help="position on the machine (or MinION serial number) to run the protocol at",
    )
    parser.add_argument(
        "--flow-cell-id",
        metavar="FLOW-CELL-ID",
        help="ID of the flow-cell on which to run the protocol. (specify this or --position)",
    )
    parser.add_argument(
        "--kit",
        required=True,
        help="Sequencing kit used with the flow-cell, eg: SQK-LSK108",
    )
    parser.add_argument(
        "--product-code",
        help="Override the product-code stored on the flow-cell and previously user-specified"
        "product-codes",
    )
    # BASECALL ARGUMENTS
    parser.add_argument(
        "--basecalling",
        action="store_true",
        help="enable base-calling using the default base-calling model",
    )
    parser.add_argument(
        "--basecall-config",
        help="specify the base-calling config and enable base-calling",
    )
    # BARCODING ARGUMENTS
    parser.add_argument(
        "--barcoding", action="store_true", help="protocol uses barcoding",
    )
    parser.add_argument(
        "--barcode-kits",
        nargs="+",
        help="bar-coding expansion kits used in the experiment",
    )
    parser.add_argument(
        "--trim-barcodes", action="store_true", help="enable bar-code trimming",
    )
    parser.add_argument(
        "--barcodes-both-ends",
        action="store_true",
        help="bar-code filtering (both ends of a strand must have a matching barcode)",
    )

    parser.add_argument(
        "--detect-mid-strand-barcodes",
        action="store_true",
        help="bar-code filtering for bar-codes in the middle of a strand",
    )
    parser.add_argument(
        "--min-score",
        type=float,
        default=0.0,
        help="read selection based on bar-code accuracy",
    )
    parser.add_argument(
        "--min-score-rear",
        type=float,
        default=0.0,
        help="read selection based on bar-code accuracy",
    )

    parser.add_argument(
        "--min-score-mid",
        type=float,
        default=0.0,
        help="read selection based on bar-code accuracy",
    )
    # ALIGNMENT ARGUMENTS
    parser.add_argument(
        "--alignment-reference",
        help="Specify alignment reference to send to basecaller for live alignment.",
    )
    parser.add_argument(
        "--bed-file", help="Specify bed file to send to basecaller.",
    )
    # Output arguments
    parser.add_argument(
        "--fastq",
        action="store_true",
        help="enables FastQ file output, defaulting to 4000 reads per file",
    )
    parser.add_argument(
        "--fastq-reads-per-file",
        type=int,
        default=4000,
        help="set the number of reads combined into one FastQ file.",
    )
    parser.add_argument(
        "--fast5",
        action="store_true",
        help="enables Fast5 file output, defaulting to 4000 reads per file, this will store raw, "
        "fastq and trace-table data",
    )
    parser.add_argument(
        "--fast5-reads-per-file",
        type=int,
        default=4000,
        help="set the number of reads combined into one Fast5 file.",
    )
    parser.add_argument(
        "--bam",
        action="store_true",
        help="enables BAM file output, defaulting to 4000 reads per file",
    )
    parser.add_argument(
        "--bam-reads-per-file",
        type=int,
        default=4000,
        help="set the number of reads combined into one BAM file.",
    )
    # Read until
    parser.add_argument(
        "--read-until-reference", type=str, help="Reference file to use in read until",
    )
    parser.add_argument(
        "--read-until-bed-file", type=str, help="Bed file to use in read until",
    )
    parser.add_argument(
        "--read-until-filter",
        type=str,
        choices=["deplete", "enrich"],
        help="Filter type to use in read until",
    )
    # Experiment
    parser.add_argument(
        "--experiment-duration",
        type=float,
        default=72,
        help="time spent sequencing (in hours)",
    )
    parser.add_argument(
        "--no-active-channel-selection",
        action="store_true",
        help="allow dynamic selection of channels to select pores for sequencing, "
        "ignored for Flongle flow-cells",
    )
    parser.add_argument(
        "--mux-scan-period",
        type=float,
        default=1.5,
        help="number of hours before a mux scan takes place, enables active-channel-selection, "
        "ignored for Flongle flow-cells",
    )
    parser.add_argument(
        "extra_args",
        metavar="ARGS",
        nargs="*",
        help="Additional arguments passed verbatim to the protocol script",
    )
    args = parser.parse_args()
    # Further argument checks
    # Read until must have a reference and a filter type, if enabled:
    if (
        args.read_until_filter is not None
        or args.read_until_reference is not None
        or args.read_until_bed_file is not None
    ):
        if args.read_until_filter is None:
            print("Unable to specify read until arguments without a filter type.")
            sys.exit(1)

        if args.read_until_reference is None:
            print("Unable to specify read until arguments without a reference type.")
            sys.exit(1)

    if args.bed_file and not args.alignment_reference:
        print("Unable to specify `--bed-file` without `--alignment-reference`.")
        sys.exit(1)

    if (args.barcoding or args.barcode_kits) and not (
        args.basecalling or args.basecall_config
    ):
        print(
            "Unable to specify `--barcoding` or `--barcode-kits` without `--basecalling`."
        )
        sys.exit(1)
    if args.alignment_reference and not (args.basecalling or args.basecall_config):
        print("Unable to specify `--alignment-reference` without `--basecalling`.")
        
        sys.exit(1)
    if not (args.fast5 or args.fastq):
        print("No output (fast5 or fastq) specified")

    return args

def is_position_selected(position, args):
    """Find if the {position} is selected by command line arguments {args}."""

    # First check for name match:
    if args.position == position.name:
        return True

    # Then verify if the flow cell matches:
    connected_position = position.connect()
    if args.flow_cell_id is not None:
        flow_cell_info = connected_position.device.get_flow_cell_info()
        if (
            flow_cell_info.user_specified_flow_cell_id == args.flow_cell_id
            or flow_cell_info.flow_cell_id == args.flow_cell_id
        ):
            return True

    return False


def startRun():
    """Entrypoint to start protocol example"""
    # Parse arguments to be passed to started protocols:
    run_id=""
    args = parse_args()
    #args = parse_args(minknowApiShellArgumentString.split())

    # Specify --verbose on the command line to get extra details about
    if args.verbose:
        logging.basicConfig(stream=sys.stdout, level=logging.DEBUG)

    # Construct a manager using the host + port provided:
    #manager = Manager(host=args.host, port=args.port, use_tls=not args.no_tls)
    manager=mkManager()
    errormessage=""
    
    # Find which positions we are going to start protocol on:
    positions = manager.flow_cell_positions()
    filtered_positions = list(
        filter(lambda pos: is_position_selected(pos, args), positions)
    )

    # At least one position needs to be selected:
    if not filtered_positions:
        errormessage="No positions selected for protocol - specify `--position` or `--flow-cell-id`"
    else:
        protocol_identifiers = {}
        for pos in filtered_positions:
            # Connect to the sequencing position:
            position_connection = pos.connect()

            # Check if a flowcell is available for sequencing
            flow_cell_info = position_connection.device.get_flow_cell_info()
            if not flow_cell_info.has_flow_cell:
                errormessage="No flow cell present in position "+str(pos)
            else:
                # Select product code:
                if args.product_code:
                    product_code = args.product_code
                else:
                    product_code = flow_cell_info.user_specified_product_code
                    if not product_code:
                        product_code = flow_cell_info.product_code

                # Find the protocol identifier for the required protocol:
                protocol_info = protocols.find_protocol(
                    position_connection,
                    product_code=product_code,
                    kit=args.kit,
                    basecalling=args.basecalling,
                    basecall_config=args.basecall_config,
                    barcoding=args.barcoding,
                    barcoding_kits=args.barcode_kits,
                )

                if not protocol_info:
                    print("Failed to find protocol for position %s" % (pos.name))
                    print("Requested protocol:")
                    print("  product-code: %s" % args.product_code)
                    print("  kit: %s" % args.kit)
                    print("  basecalling: %s" % args.basecalling)
                    print("  basecall_config: %s" % args.basecall_config)
                    print("  barcode-kits: %s" % args.barcode_kits)
                    print("  barcoding: %s" % args.barcoding)
                    errormessage="Protocol build error, consult application log."
                else:
                    # Store the identifier for later:
                    protocol_identifiers[pos.name] = protocol_info.identifier

                    # Start protocol on the requested postitions:
                    print("Starting protocol on %s positions" % len(filtered_positions))
                    for pos in filtered_positions:

                        # Connect to the sequencing position:
                        position_connection = pos.connect()

                        # Find the protocol identifier for the required protocol:
                        protocol_identifier = protocol_identifiers[pos.name]

                        # Now select which arguments to pass to start protocol:
                        print("Starting protocol %s on position %s" % (protocol_identifier, pos.name))

                        # Set up user specified product code if requested:
                        if args.product_code:
                            position_connection.device.set_user_specified_product_code(
                                code=args.product_code
                            )

                        # Build arguments for starting protocol:
                        basecalling_args = None
                        if args.basecalling or args.basecall_config:
                            barcoding_args = None
                            alignment_args = None
                            if args.barcode_kits or args.barcoding:
                                barcoding_args = protocols.BarcodingArgs(
                                    args.barcode_kits,
                                    args.trim_barcodes,
                                    args.barcodes_both_ends,
                                    args.detect_mid_strand_barcodes,
                                    args.min_score,
                                    args.min_score_rear,
                                    args.min_score_mid,
                                )

                            if args.alignment_reference:
                                alignment_args = protocols.AlignmentArgs(
                                    reference_files=[args.alignment_reference], bed_file=args.bed_file,
                                )

                            basecalling_args = protocols.BasecallingArgs(
                                config=args.basecall_config,
                                barcoding=barcoding_args,
                                alignment=alignment_args,
                            )

                        read_until_args = None
                        if args.read_until_filter:
                            read_until_args = protocols.ReadUntilArgs(
                                filter_type=args.read_until_filter,
                                reference_files=[args.read_until_reference],
                                bed_file=args.read_until_bed_file,
                                first_channel=None,  # These default to all channels.
                                last_channel=None,
                            )

                        def build_output_arguments(args, name):
                            if not getattr(args, name):
                                return None
                            return protocols.OutputArgs(
                                reads_per_file=getattr(args, "%s_reads_per_file" % name)
                            )

                        fastq_arguments = build_output_arguments(args, "fastq")
                        fast5_arguments = build_output_arguments(args, "fast5")
                        bam_arguments = build_output_arguments(args, "bam")

                        # print the protocol parameters
                        print("position_connection "+str(position_connection))
                        print("protocol_identifier "+str(protocol_identifier))
                        print("args.sample_id "+str(args.sample_id))
                        print("args.experiment_group "+str(args.experiment_group))
                        print("basecalling_args "+str(basecalling_args)) 
                        print("read_until_args "+str(read_until_args))
                        print("fastq_arguments "+str(fastq_arguments)) #fastq_arguments OutputArgs(reads_per_file=400)
                        print("fast5_arguments "+str(fast5_arguments)) #fast5_arguments OutputArgs(reads_per_file=400)
                        print("bam_arguments "+str(bam_arguments))
                        print("args.no_active_channel_selection"+str(args.no_active_channel_selection))
                        print("args.mux_scan_period"+str(args.mux_scan_period))
                        print("args.experiment_duration "+str(args.experiment_duration))
                        print("args.extra_args "+str(args.extra_args))  # Any extra args passed.

                        # Now start the protocol:
                        run_id = protocols.start_protocol(
                            position_connection,
                            protocol_identifier,
                            sample_id=args.sample_id,
                            experiment_group=args.experiment_group,
                            basecalling=basecalling_args,
                            read_until=read_until_args,
                            fastq_arguments=fastq_arguments,
                            fast5_arguments=fast5_arguments,
                            bam_arguments=bam_arguments,
                            disable_active_channel_selection=args.no_active_channel_selection,
                            mux_scan_period=args.mux_scan_period,
                            experiment_duration=args.experiment_duration,
                            args=args.extra_args,  # Any extra args passed.
                        )

                        #print("Started protocol %s" % run_id)
    return errormessage+run_id # one of them should be ""


# In[ ]:


def stopRun(minionId): # stop an existing run (if any) for a MinION device
    manager=mkManager()
    positions = list(manager.flow_cell_positions())
    filtered_positions = list(filter(lambda pos: pos.name == minionId, positions))
    # Connect to the grpc port for the position:
    connection = filtered_positions[0].connect()
    protocols = connection.protocol.list_protocol_runs()
    bufferedRunIds = protocols.run_ids
    thisMessage="No protocol running, nothing was stopped."
    c=0
    for b in bufferedRunIds:
        try:
            connection.protocol.stop_protocol()
            thisMessage="Protocol "+b+" stopped on "+minionId+"."
        except:
            c=c+1
    return thisMessage


# In[ ]:


# from minknow_api demos, start_seq.py
def is_position_selected(position, args):
    """Find if the {position} is selected by command line arguments {args}."""
    if args.position == position.name: # First check for name match:
        return True
    connected_position = position.connect()  # Then verify if the flow cell matches:
    if args.flow_cell_id is not None:
        flow_cell_info = connected_position.device.get_flow_cell_info()
        if (flow_cell_info.user_specified_flow_cell_id == args.flow_cell_id
            or flow_cell_info.flow_cell_id == args.flow_cell_id):
            return True
    return False


# In[ ]:


def getMinKnowApiStatus(deviceString): # MinKNOW status per device
    replyString=""
    testHost="localhost"
    manager=mkManager()
    positions = list(manager.flow_cell_positions())
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position
    # determine if anything is running and the kind of run, via set temperature
    replyString=replyString+"acquisition.get_acquisition_info().state: "+str(connection.acquisition.get_acquisition_info().state)+"<br>"
    replyString=replyString+"acquisition.current_status(): "+str(connection.acquisition.current_status())+"<br>"
    replyString=replyString+"minion_device.get_settings().temperature_target.min: "+str(connection.minion_device.get_settings().temperature_target.min)+"<br>"
    replyString=replyString+"device.get_temperature(): " + str(connection.device.get_temperature().minion.heatsink_temperature)+"<br>"
    replyString=replyString+"device.get_bias_voltage(): " + str(connection.device.get_bias_voltage())+"<br>"
    return replyString


# In[ ]:


def getActiveRun(deviceString):
    manager=mkManager()
    positions = list(manager.flow_cell_positions())
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position
    try:
        activeRun=connection.acquisition.get_current_acquisition_run().run_id # error if no acquisition is running, same as with acquisitio.current_status(), no acquisition until temperature reached
    except:
        activeRun="none"
    return activeRun


# In[ ]:


def getRealDeviceActivity(deviceString):            # seq. runs: 34 degC and flow cell checks 37 degC target 
    manager=mkManager()                             # temperatures seem to be the only way to determine if
    positions = list(manager.flow_cell_positions()) # a device has been started
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position
    targetTemp=str(connection.minion_device.get_settings().temperature_target.min)
    returnValue=""
    if targetTemp=="34.0":
        returnValue="sequencing"
    elif targetTemp=="37.0":
        returnValue="checking flow cell"
    elif targetTemp=="35.0":
        returnValue="idle"
    return returnValue


# In[ ]:


def getThisRunState(deviceString): # obtain further information about a particular device / run
    manager=mkManager()
    positions = list(manager.flow_cell_positions())
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position
    try:
        thisRunState="Run state for "+deviceString+": "
        thisRunState=thisRunState+str(connection.protocol.get_current_protocol_run().state)+"/"
        thisRunState=thisRunState+str(connection.acquisition.get_acquisition_info().state)
    except:
        thisRunState="No state information in MinKNOW buffer for "+deviceString
    return thisRunState


# In[ ]:


def getThisRunSampleID(deviceString): # get SampleID from MinKNOW by device, only available after data
    manager=mkManager()               # acquisition as been initiated by MinKNOW.
    positions = list(manager.flow_cell_positions())
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position
    try:
        thisRunSampleID=connection.protocol.get_current_protocol_run().user_info.sample_id.value
    except:
        thisRunSampleID="No sampleId information in MinKNOW buffer for "+deviceString
    return thisRunSampleID   


# In[ ]:


def getThisRunYield(deviceString): # get run yield by device. The data of the previous run will remain 
    manager=mkManager()            # in the buffer until acquisition (not just a start) of a new run
    positions = list(manager.flow_cell_positions()) # have been initiated.
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position
    try:
        acqinfo=connection.acquisition.get_acquisition_info()
        thisRunYield="Run yield for "+deviceString+"("+acqinfo.run_id+"):&nbsp;"
        thisRunYield=thisRunYield+str(acqinfo.yield_summary)
    except:
        thisRunYield="No yield information in MinKNOW buffer for "+deviceString
    return thisRunYield


# In[ ]:


def getThisRunOutput(deviceString,sampleName,runId): # get run yield by device, sampleName, runId
    thisRunOutput=[-1,-1] # defaults in case of error / missing information
    manager=mkManager()            # in the buffer until acquisition (not just a start) of a new run
    positions = list(manager.flow_cell_positions()) # have been initiated.
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position
    readCount=-3
    calledBases=-3
    if getThisRunSampleID(deviceString)==sampleName: # check that runID and sampleID match
        readCount=-4
        calledBases=-4
        if connection.acquisition.get_current_acquisition_run().run_id==runId:
            if connection.acquisition.current_status()!="status: READY": # i.e., working
                try:
                    acq=connection.acquisition.get_acquisition_info()
                    readCount=acq.yield_summary.basecalled_pass_read_count
                    calledBases=acq.yield_summary.basecalled_pass_bases
                except:
                    readCount=-5
                    calledBases=-5
    thisRunOutput=[readCount,calledBases]
    return thisRunOutput # shall be a list


# In[ ]:


def getThisRunEstimatedOutput(deviceString,sampleName,runId): # get run yield by device, sampleName, runId
    thisRunOutput=[-1,-1] # defaults in case of error / missing information
    manager=mkManager()            # in the buffer until acquisition (not just a start) of a new run
    positions = list(manager.flow_cell_positions()) # have been initiated.
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position
    readCount=-3
    calledBases=-3
    if getThisRunSampleID(deviceString)==sampleName: # check that runID and sampleID match
        readCount=-4
        calledBases=-4
        if connection.acquisition.get_current_acquisition_run().run_id==runId:
            if connection.acquisition.current_status()!="status: READY": # i.e., working
                try:
                    acq=connection.acquisition.get_acquisition_info()
                    readCount=acq.yield_summary.basecalled_pass_read_count
                    calledBases=acq.yield_summary.estimated_selected_bases
                except:
                    readCount=-5
                    calledBases=-5
    thisRunOutput=[readCount,calledBases]
    return thisRunOutput # shall be a list


# In[ ]:


def getThisRunInformation(deviceString): # get current run information. Only available after data acquisition
    manager=mkManager()                  # has started.
    positions = list(manager.flow_cell_positions())
    filtered_positions = list(filter(lambda pos: pos.name == deviceString, positions))
    connection = filtered_positions[0].connect() # Connect to the grpc port for the position    
    try:
        thisRunInfo="Run information for "+deviceString+"<br><br>"+str(connection.protocol.get_current_protocol_run())
    except:
        thisRunInfo="No protocol information in MinKNOW buffer for "+deviceString
    return thisRunInfo


# In[ ]:


def thisRunWatcherTerminator(deviceString,sampleName):
    realRunId=getActiveRun(deviceString) #
    currentBases=getThisRunEstimatedOutput(deviceString,sampleName,realRunId)[1]
    currentBasesString=str(round(currentBases/1e6,2))
    wantedBasesString=str(round(wantedBases/1e6,2))
    myString="<html><head>"
    myString=myString+"<title>"+currentBasesString+"/"+wantedBasesString+"MB:"+sampleName+"</title>"
    if currentBases < wantedBases: # don't refresh after showing the STOP state
        myString=myString+"<meta http-equiv='refresh' content='10'>"
    myString=myString+"</head><body>"
    myString=myString+"<b>Automatic run terminator</b> for sample <b>"+sampleName+ "</b>, run ID="+realRunId+" on "+deviceString+" when reaching "+wantedBasesString+" MB, now "+currentBasesString+" MB"
    myString=myString+"<hr>"
    myString=myString+"Last refresh at "+datetimestringnow()+".<hr>"
    if currentBases > wantedBases:
        stopRun(deviceString)
        myString=myString+"STOPPED at "+datetimestringnow()
    elif currentBases==0:
        myString=myString+"IDLE / MUX / ETC"
    else:
        myString=myString+"RUNNING"
    myString=myString+"</body></html>"
    return myString


# ### 3. CNV Plotter

# In[ ]:


def createCNVPlot(sampleName): # create a genome-wide copy number plot (all-in-one function)
    with tqdm(total=6) as cnvpBar:
        cnvpBar.set_description('CNVP('+sampleName+'), loading reference data')
        cnvpBar.update(1)
        startTime = datetime.datetime.now()
        runPath=nanodipOutputDir+"/"+sampleName
        ChromOffsets = pandas.read_csv(chrLengthsFile, delimiter='\t', header=None, index_col=0)
        validChromosomes=list(ChromOffsets.index)
        ChromOffsetCenters=[]
        for c in range(0,len(validChromosomes)-1):
            ChromOffsetCenters.append((ChromOffsets[2][c]+ChromOffsets[2][c+1])/2)
        lastChromosome=len(validChromosomes)-1
        ChromOffsetCenters.append((ChromOffsets[2][lastChromosome]+ChromOffsets[2][lastChromosome]+ChromOffsets[1][lastChromosome])/2) # last chromosome
        centromereLocations = pandas.read_csv(centromereLocationsBed, delimiter='\t', header=None, index_col=0)
        centromereLocations.loc['chr1'][1]
        centromereOffsets = []
        for c in validChromosomes:
            centromereOffsets.append(ChromOffsets.loc[c][2] + centromereLocations.loc[c][1])
        cnvpBar.set_description('CNVP('+sampleName+'), loading nanopore data')
        cnvpBar.update(1)
        bamFiles=[] # find all bam files
        for root, dirnames, filenames in os.walk(runPath):
            for filename in fnmatch.filter(filenames, '*.bam'):
                bamFiles.append(os.path.join(root, filename))
        cnvScatter=[]
        for thisBam in bamFiles:
            samfile = pysam.AlignmentFile(thisBam, "rb") # pysam coordinates start with 0 while samtools starts with 1 ! See https://pysam.readthedocs.io/en/latest/faq.html#pysam-coordinates-are-wrong
            for thisChromosome in validChromosomes:
                thisChromOffset=int(ChromOffsets.loc[[thisChromosome]][2])
                for read in samfile.fetch(thisChromosome):
                    cnvScatter.append(read.pos+thisChromOffset)
        #print("Number of reads:"+str(len(cnvScatter)))
        #print(max(cnvScatter)/len(cnvScatter))
        cnvpBar.set_description('CNVP('+sampleName+'), determining bin size')
        cnvpBar.update(1)
        binwidth=30*max(cnvScatter)/len(cnvScatter)
        figure(figsize=(20, 3), dpi=120)
        xy=plt.hist(cnvScatter, bins=numpy.arange(min(cnvScatter), max(cnvScatter) + binwidth, binwidth),color="k")
        #plt.vlines(ChromOffsets[2], 0,numpy.max(xy[0]), colors='c', linestyles='solid', label='')
        #plt.title(sampleName, fontdict=None, loc='center', pad=None)
        #plt.yscale('log')
        cnvpBar.set_description('CNVP('+sampleName+'), cleaning data')
        cnvpBar.update(1)
        plotX=xy[1][0:len(xy[0])] # exclude regions with no mapped reads from plot
        plotY=xy[0]
        cleanPlotX=[]
        cleanPlotY=[]
        for p in range(0,len(plotY)):
            if plotY[p]>0:
                cleanPlotX.append(plotX[p])
                cleanPlotY.append(plotY[p])
        cleanPlotX=numpy.array(cleanPlotX) # convert back to numpy array (required for numpy functions)
        cleanPlotY=numpy.array(cleanPlotY)
        yStd=numpy.std(plotY)
        yMean=numpy.mean(plotY)
        yMedian=numpy.median(plotY)
        cleanCoarseX=[]     # local means, cleaned
        cleanCoarseY=[]
        localBinSize=int(10e6)
        localBinStep=int(0.5e6)
        halfLocalBinSize=localBinSize/2
        cnvpBar.set_description('CNVP('+sampleName+'), plotting data')
        cnvpBar.update(1)
        for x in range(0,int(numpy.round(numpy.max(cleanPlotX))),localBinStep):
            thisSlice=cleanPlotY[numpy.logical_and(cleanPlotX >= x,  cleanPlotX <= x+localBinSize)]
            if len(thisSlice)>0:
                cleanCoarseX.append(numpy.median([x,x+localBinSize]))
                cleanCoarseY.append(numpy.median(thisSlice))
        cleanYMedian=numpy.median(cleanPlotY)
        cleanYLower=numpy.min(cleanPlotY)
        cleanYUpper=yMedian*2
        matplotlib.use('Agg')
        figure(figsize=(20, 6), dpi=120)
        plt.ylim(cleanYLower,cleanYUpper)
        plt.scatter(cleanPlotX,cleanPlotY,s=0.2,color='gray',linewidths=1)
        plt.scatter(cleanCoarseX,cleanCoarseY,s=1,linewidths=5,c=cleanCoarseY,cmap=plt.cm.coolwarm_r,vmin=cleanYLower,vmax=cleanYUpper)
        plt.hlines(yMedian, 0, max(cleanPlotX), colors='gray', linestyles='solid', label='') # median line
        plt.vlines(ChromOffsets[2], cleanYLower, cleanYUpper, colors='gray', linestyles='solid', label='')
        plt.vlines(ChromOffsets[2][len(ChromOffsets[2])-1]+ChromOffsets[1][len(ChromOffsets[2])-1], cleanYLower, cleanYUpper, colors='gray', linestyles='solid', label='') # terminating vline
        plt.vlines(centromereOffsets, cleanYLower, cleanYUpper, colors='gray', linestyles='dashed', label='')
        plt.title("Sample ID: "+sampleName, fontdict=None, loc='center', pad=None)
        plt.xlabel('Number of mapped reads: '+str(len(cnvScatter)))
        plt.ylabel('reads per '+ str(round(binwidth/1e6*100)/100) +' MB bin')
        plt.xticks(ChromOffsetCenters, validChromosomes, rotation=90)
        plt.savefig(nanodipReportDir+'/'+sampleName+'_CNVplot.png', bbox_inches='tight')
        readCountFile = open(nanodipReportDir+'/'+sampleName+'_alignedreads.txt',"w")
        readCountFile.write(str(len(cnvScatter)))
        readCountFile.close()
        logpr(verbosity,"CNVP end")
        cnvpBar.set_description('CNVP('+sampleName+'), done')
        cnvpBar.update(1)
        endTime = datetime.datetime.now()
        logpr(verbosity,"Start: "+str(startTime))
        logpr(verbosity,"End  : "+str(endTime))
        logpr(verbosity,"Dur. : "+str(endTime-startTime))


# ### 4. UMAP Methylation Plotter

# In[ ]:


def methylationUMAP(sampleName,referenceName): # create a methylation UMAP plot (all-in-one function)
    import umap
    startTime = datetime.datetime.now()
    logpr(verbosity,"UMAP Plot initiated for "+sampleName)
    with tqdm(total=8) as umapBar:
        umapBar.set_description('UMAP('+sampleName+'), loading annotation')
        umapBar.update(1)
        binFiles=listdir(binDir) # collect reference case binary file names
        referenceString=referenceName.replace(".xlsx","")
        referenceSheetFile=referenceDir+"/"+referenceName # load reference annotation
        referenceSheet=openpyxl.load_workbook(referenceSheetFile)
        referenceList = referenceSheet.active
        col_names = []
        sentrixID  = referenceList['A']
        methClass  = referenceList['B']
        customText = referenceList['C']
        for x in range(3): 
            logpr(verbosity,sentrixID[x].value)
            logpr(verbosity,methClass[x].value)
            logpr(verbosity,customText[x].value)
        indexFile=open(binIndex, "r") # load CpG site index file (contains index for methylation float binary data)
        indexCol=indexFile.read().split("\n")
        indexFile.close()
        umapBar.set_description('UMAP('+sampleName+'), loading and processing methylation data from Nanopore run')
        umapBar.update(1)
        logpr(verbosity,len(indexCol))
        methoverlapPath=nanodipOutputDir+"/"+sampleName # collect matching CpGs from sample
        methoverlapTsvFiles=[] # find all *methoverlap.tsv files
        for root, dirnames, filenames in os.walk(methoverlapPath):
            for filename in fnmatch.filter(filenames, '*methoverlap.tsv'):
                methoverlapTsvFiles.append(os.path.join(root, filename))
        methoverlap=[]
        first=True
        for f in methoverlapTsvFiles:
            try: # some fast5 files do not contain any CpGs
                m = pandas.read_csv(f, delimiter='\t', header=None, index_col=0)
                if first:
                    methoverlap = m
                    first = False
                else:
                    methoverlap = methoverlap.append(m)
            except:
                logpr(verbosity,"empty file encountered, skipping")
        if len(methoverlap)>0:
            logpr(verbosity,str("Number of 450K overlap CpGs: "+str(len(methoverlap))))
            overlapProbes=methoverlap.index
            existingProbes=set(indexCol).intersection(overlapProbes) # some probes have been skipped from the reference set, e.g. sex chromosomes
            matching = [indexCol.index(i) for i in existingProbes]
            logpr(verbosity,"overlap probes in cleaned reference data: "+str(len(matching)))
            fileNumbers = []
            binSuffix="_betas_filtered.bin"
            missingFiles=[] # determine if there are entries in the annotation without corresponding methylation binary file
            c=0
            for s in sentrixID:
                try:
                    fileNumbers.append(binFiles.index(s.value+binSuffix))
                except: # e.g. file not available
                    missingFiles.append(c)
                c=c+1
            logpr(verbosity,fileNumbers)
            betaValues=numpy.full([len(matching),len(fileNumbers)],-1, dtype=float, order='C') # create an empty array with -1
            logpr(verbosity,betaValues)
            umapBar.set_description('UMAP('+sampleName+'), loading overlap CpGs from reference data')
            umapBar.update(1)
            matchJumps=numpy.full([len(matching)],-1, dtype=int, order='C')
            matchJumps[0]=matching[0] # create jump list for binary file, add first entry
            if len(matching)>1:
                for m in range(1,len(matching)): # create jump distances for binary file
                    matchJumps[m]=matching[m]-matching[m-1]-1 # concatenate to list
            logpr(verbosity,len(matchJumps))
            betaValues = [ [ None for y in range( len(matching) ) ] for x in range( 1 ) ]
            p_bar = tqdm(range(len(fileNumbers))) # progress bar (development only)
            for f in p_bar:
                betasFilename=binDir+"/"+binFiles[fileNumbers[f]]
                with open(betasFilename, 'rb') as betasFile:
                    allBetaSingleFile = numpy.fromfile(betasFile, dtype=float) # read float with numpy into regular python array (faster) 
                    allBetaSingleFile = numpy.digitize(allBetaSingleFile,bins=[methylCutOff])
                    betaValues.append(allBetaSingleFile[numpy.array(matching)])
                    p_bar.set_description('UMAP('+sampleName+'), loading ref. dataset no. '+str(f))
                betasFile.close()
            umapBar.set_description('UMAP('+sampleName+'), merging reference and nanopore data')
            umapBar.update(1)
            betaValues = numpy.array(betaValues)
            betaValues = numpy.delete(betaValues, 0, 0)
            methoverlapNum=methoverlap.to_numpy()
            diagnosticCaseCgs=[]
            methoverlapCgnames=methoverlap.loc[existingProbes].index # deterine overlap CpG names
            for i in existingProbes:
                thisCg=numpy.mean(methoverlap.loc[[i]].values)
                diagnosticCaseCgs.append(thisCg)
            thisDiagnosticCase=numpy.digitize(diagnosticCaseCgs,bins=[methylCutOff]) # append the nanopore case
            betaValues2=numpy.vstack([betaValues, thisDiagnosticCase]) # convert to numpy array for UMAP function
            del betaValues # free memory
            umapBar.set_description('UMAP('+sampleName+'), calculating embedding')
            umapBar.update(1)
            embeddingAll = umap.UMAP().fit_transform(betaValues2[:,]) # generate UMAP plot
            logpr(verbosity,"\n"+str(embeddingAll))
            umapBar.set_description('UMAP('+sampleName+'), plotting UMAP')
            umapBar.update(1)
            l=len(embeddingAll)-1  # get UMAP coordinates of nanopore case (i.e., last entry in array)
            nanoX=embeddingAll[l,0]
            nanoY=embeddingAll[l,1]
            selectedSentrixIds = [ binFiles[i] for i in fileNumbers]
            logpr(verbosity,len(selectedSentrixIds))
            selectedSentrixIds.append(sampleName)
            logpr(verbosity,len(selectedSentrixIds))
            annoList=[] # create an annotation list and append nanopore case as the last entry
            c=0
            for mc in methClass:
                if c not in missingFiles:
                    annoList.append(mc.value)
                c=c+1
            annoList.append(sampleName)
            embeddingAll=numpy.array(embeddingAll) # convert UMAP data to numpy array
            numberRef=str(len(embeddingAll))+" ref. cases"
            numberCpG=str(len(methoverlap))+" CpGs"
            umapTitle="UMAP for "+sampleName+" against "+referenceName+", "+numberRef+", "+numberCpG
            logpr(verbosity,type(embeddingAll))
            logpr(verbosity,embeddingAll.shape)
            logpr(verbosity,embeddingAll)
            fig2 = px.scatter(x=embeddingAll[:,0], # create UMAP figure with all cases
                              y=embeddingAll[:,1],
                              labels={"x":"UMAP 0",
                                      "y":"UMAP 1"},
                              title=umapTitle, 
                              color=annoList, 
                              hover_name=selectedSentrixIds,
                              render_mode=plotlyRenderMode) #
            fig2.add_annotation(x=nanoX, y=nanoY,
                                text=sampleName,
                                showarrow=True,
                                arrowhead=1)
            fig2.update_yaxes(scaleanchor = "x", scaleratio = 1)
            outPlot=nanodipReportDir+"/"+sampleName+"_"+referenceString+"_UMAP_all.html" # write to HTML file
            fig2.write_html(outPlot)
            fig2.write_image(nanodipReportDir+"/"+sampleName+"_"+referenceString+"_UMAP_all.png")    # plotly image export requires kaleido, install with pip install -U kaleido; needs reloading of plotly to take effect
            umapBar.set_description('UMAP('+sampleName+'), calculating distance ranking')
            umapBar.update(1)
            distances = [] # create distance ranking
            sentrixList = []
            c=0
            for s in sentrixID:
                if c not in missingFiles:
                    sentrixList.append(s.value)
                c=c+1
            sentrixList.append("thisCase")
            mcList = []
            c=0
            for s in methClass:
                if c not in missingFiles:
                    mcList.append(s.value)
                c=c+1
            mcList.append("thisCase")
            txtList = []
            c=0
            for s in methClass:
                if c not in missingFiles:
                    txtList.append(s.value)
                c=c+1
            txtList.append("thisCase")
            caseX=embeddingAll[len(embeddingAll)-1,0]
            caseY=embeddingAll[len(embeddingAll)-1,1]
            xList = []
            yList = []
            for c in embeddingAll:
                distances.append(numpy.sqrt(numpy.power(caseX-c[0],2)+numpy.power(caseY-c[1],2))) # calculate distance
                xList.append(c[0])
                yList.append(c[1])
            distanceRanking = pandas.DataFrame({'distance':distances,'methClass':mcList,'txt':txtList,
                                                'sentrix_ID':sentrixList,'X':xList,'Y':yList})
            distanceRanking = distanceRanking.sort_values(by='distance', axis=0, ascending=True, inplace=False, kind='quicksort')
            # distanceRanking[0:20]
            wb = openpyxl.Workbook()     # write plot coordinates to xlsx
            ws = wb.active # grab the active worksheet
            ws['A1'] = "Sentrix_ID"     # Data can be assigned directly to cells
            ws['B1'] = "X"     
            ws['C1'] = "Y"
            for thisRow in range(len(embeddingAll)):     # Rows can also be appended
                ws.append([selectedSentrixIds[thisRow], embeddingAll[thisRow][0], embeddingAll[thisRow][1]])
            wb.save(nanodipReportDir+"/"+sampleName+"_UMAP.xlsx")     # Save the file
            closeupDf=distanceRanking[0:topMatch] # generate plot of thisCase surrounding reference datasets
            closeupList=closeupDf.values.tolist()
            markerSize=5 # marker size for plotly
            fig3=px.scatter(x=closeupDf["X"],
                            y=closeupDf["Y"],
                            labels={"x":"UMAP 0",
                                    "y":"UMAP 1"},
                            hover_name=closeupDf["sentrix_ID"],
                            title="Close-up "+umapTitle,
                            color=closeupDf["methClass"],
                            render_mode=plotlyRenderMode)
            fig3.update_traces(marker=dict(size=markerSize))
            fig3.add_annotation(x=nanoX, y=nanoY, text=sampleName, showarrow=True, arrowhead=1)
            for ds in closeupDf.values.tolist(): # add transparent hyperlinks to reference CNV plots (e.g., on public EpiDiP.org server)
                fig3.add_annotation(x=ds[4], y=ds[5],
                                    text="<a href='"+cnvLinkPrefix+ds[3]+cnvLinkSuffix+"' target='_blank'>&nbsp;</a>",
                                    showarrow=False, arrowhead=1)
            topRadius=closeupList[len(closeupList)-1][0]
            fig3.add_shape(type="circle",
               x0=nanoX-topRadius,
               y0=nanoY-topRadius,
               x1=nanoX+topRadius,
               y1=nanoY+topRadius,
               line_color="black",
               line_width=0.5)
            fig3.update_yaxes(scaleanchor = "x", scaleratio = 1)
            outPlot=nanodipReportDir+"/"+sampleName+"_"+referenceString+"_UMAP_top.html"
            fig3.write_html(outPlot)
            fig3.write_image(nanodipReportDir+"/"+sampleName+"_"+referenceString+"_UMAP_top.png")
            # create PDF-compatible HTML table with proper cell padding, no borders etc.
            htmlTable="<table border=0>"
            htmlTable+="<tr><td><b>methClass</b></td><td><b>distance</b></td><td><b>txt</b></td><td><b>sentrix_ID</b></td></tr>"
            for i in closeupList:
                htmlTable+="<tr><td>"+str(i[1])+"</td><td>"+str(i[0])+"</td><td>"+str(i[2])+"</td><td>"+str(i[3])+"</td></tr>"
            # generate PDF report
            nanodipReport="<html><head><title>"+sampleName+"</title><body><h1>"+sampleName+"</h1>"+sampleName+"<br>"+htmlTable+"</body>"
            convert_html_to_pdf(nanodipReport, nanodipReportDir+"/"+sampleName+"_"+referenceString+"_NanoDiP_ranking.pdf")
            cpgCountFile = open(nanodipReportDir+'/'+sampleName+'_cpgcount.txt',"w")
            cpgCountFile.write(str(len(methoverlap)))
            cpgCountFile.close()
            logpr(verbosity,"UMAP end")
            endTime = datetime.datetime.now()
            logpr(verbosity,"Start: "+str(startTime))
            logpr(verbosity,"End  : "+str(endTime))
            logpr(verbosity,"Dur. : "+str(endTime-startTime))
            umapBar.set_description('UMAP('+sampleName+'), completed')
            umapBar.update(1)
        else:
            outPlot=nanodipReportDir+"/"+sampleName+"_"+referenceString+"_UMAP_all.html"
            with open(outPlot, 'w') as txtfile:
                txtfile.write("<html><body>No data to plot.</body></html>")


# ### 5. Report Generator

# In[ ]:


def generatePdfReport(sampleId,referenceId):
    referenceName=referenceId.replace(".xlsx","")
    runDate=datetimestringnow()
    runSystemName=str(socket.gethostname())
    umapTopPlotPath=nanodipReportDir+"/"+sampleId+"_"+referenceName+"_UMAP_top.png"
    cnvPlotPath=nanodipReportDir+"/"+sampleId+"_CNVplot.png"
    readCountPath=nanodipReportDir+"/"+sampleId+"_alignedreads.txt"
    cpgCountPath=nanodipReportDir+"/"+sampleId+"_cpgcount.txt"
    reportPdfName=sampleId+"_"+referenceName+"_NanoDiP_report.pdf"
    reportPdfPath=nanodipReportDir+"/"+reportPdfName
    detectedBarcode=getPredominantBarcode(sampleId)
    readCount=open(readCountPath,"r")
    validReads=str(readCount.read())
    readCount.close()
    cpgCount=open(cpgCountPath,"r")
    overlapcpgCount=str(cpgCount.read())
    cpgCount.close()
    htmlcode="""
        <html><body>
        <h1><img src='"""+epidipLogoPath+"""' width='30' align='top'>&nbsp;&nbsp;
        NanoDiP Report for Sample """+sampleId+"""</h1>
        <table border='0'>
        <tr><td width='100'>Generated on</td><td width='10'>:</td><td>"""+runSystemName+""" / """+runDate+"""</td></tr>
        <tr><td>Detected barcode</td><td>:</td><td>"""+detectedBarcode+"""</td></tr>
        <tr><td>Valid reads</td><td>:</td><td>"""+validReads+"""</td></tr>
        <tr><td>450K overlap CpG count</td><td>:</td><td>"""+overlapcpgCount+"""</td></tr>
        <tr><td>Reference data</td><td>:</td><td>"""+referenceName+"""</td></tr>
        </table>
        <h2>Methylation UMAP plot</h2>
        <img src='"""+umapTopPlotPath+"""' width='500'><br>
        <h2>Copy number plot</h2>
        <img src='"""+cnvPlotPath+"""' width='700'>
        </body><html>
    """
    convert_html_to_pdf(htmlcode, reportPdfPath)
    return "<html><a href='reports/"+reportPdfName+"'>"+"PDF report generated for "+sampleId+", click this link to open it.</a></html>"


# ### 6. User Interface Functions

# In[ ]:


def systemStats(): # obtain generic system parameters
    total, used, free = shutil.disk_usage(minknowDataDir)
    m="<tt><b>SSD or HDD usage</b><br>"
    m=m+"Total: "+str(total // (2**30))+" GB<br>"
    m=m+"Used : "+str(used // (2**30))+" GB<br>"
    m=m+"Free : "+str(free // (2**30))+" GB<br>"
    m=m+"<br><b>Memory</b><br>"
    m=m+"free : "+str(round(psutil.virtual_memory().available * 100 / psutil.virtual_memory().total))+"%<br>"
    m=m+"<br><b>CPU: </b><br>"    
    m=m+"usage: "+str(round(psutil.cpu_percent()))+"%<br>"
    m=m+"CpG&nbsp;&nbsp;active runs: "+str(MinKnowIfPApi.cpgQueue)+" <a href='resetQueue?queueName=cpg'>reset queue</a><br>"
    m=m+"CNVP active runs: "+str(MinKnowIfPApi.cnvpQueue)+" <a href='resetQueue?queueName=cnvp'>reset queue</a><br>"
    m=m+"UMAP active runs: "+str(MinKnowIfPApi.umapQueue)+" <a href='resetQueue?queueName=umap'>reset queue</a><br>"
    m=m+"<br><br>"
    m=m+"<a href='restartServer'>Restart NanoDiP</a> - <font color='#FF0000'><b>USE WITH CAUTION!</b></font>"
    m=m+"</tt>" 
    return m


# In[ ]:


def analysisLaunchTable(): # present a table from which analyses can be started in a post-hoc manner
    allowedRunFolders=[]
    for a in getRuns():
        for b in analysisExclude:
            if not b in a:
                allowedRunFolders.append(a)
    annotations=getReferenceAnnotations()
    lt="<tt><font size='-2'><table border=1>" # lt = launch table, a HTML table with preformed commands to launch analyses
    lt=lt+"<thead><tr><th align='left'><b>Sample ID</b></th><th align='left'><b>CpGs</b></th><th align='left'><b>CNV</b></th>"
    for a in annotations:
        lt=lt+"<th align='left'><b>UMAP against<br>"+a.replace(".xlsx", "")+"</b></th>"
    lt=lt+"</tr></thead><tbody>"
    for r in range(len(allowedRunFolders)):
        lt=lt+"<tr>"
        lt=lt+"<td>"+allowedRunFolders[r]+"</td>"
        lt=lt+"<td><a href='./analysisLauncher?functionName=methylationPoller&sampleName="+allowedRunFolders[r]+"&refAnno=None' target='_blank' rel='noopener noreferrer' title='"+allowedRunFolders[r]+": CpGs'>get CpGs</a></td>"
        lt=lt+"<td><a href='./analysisLauncher?functionName=cnvplot&sampleName="+allowedRunFolders[r]+"&refAnno=None' target='_blank' rel='noopener noreferrer' title='"+allowedRunFolders[r]+": CNV'>plot CNV</a></td>"
        for a in annotations:
            lt=lt+"<td><a href='./analysisLauncher?functionName=umapplot&sampleName="+allowedRunFolders[r]+"&refAnno="+a+"' target='_blank' rel='noopener noreferrer' title='"+allowedRunFolders[r]+": "+a.replace(".xlsx", "")+"'>plot UMAP</a>&nbsp;"
            lt=lt+"<a href='./makePdf?sampleName="+allowedRunFolders[r]+"&refAnno="+a+"' target='_blank' rel='noopener noreferrer' title='"+allowedRunFolders[r]+": "+a.replace(".xlsx", "")+"'>make PDF</a></td>"    
        lt=lt+"</tr>"
    lt=lt+"</tbody></table></font></tt>"
    return(lt)


# In[ ]:


def listRunsTable(): # return all run folders as HTML table
    runFoldersHtml="<table border=1>"
    for r in getRuns():
        runFoldersHtml=runFoldersHtml+"<tr><td>"+r+"</td></tr>"
    runFoldersHtml=runFoldersHtml+"</table>"
    return(runFoldersHtml)


# In[ ]:


def collectPastAnalyses(): # list all analysis result files
    fl=[] # empty file file
    for f in listdir(nanodipReportDir):
        for s in resultEndings:
            if s in f:
                fl.append([f,float(os.path.getmtime(nanodipReportDir+"/"+f))])
    fl.sort(key=lambda row: (row[1], row[0]), reverse=True) # sort based on modif. date
    fl=[j.pop(0) for j in fl] # remove date column after sorting
    return fl


# In[ ]:


def makePastAnalysesTable(): # create an HTML table displaying links to completed analysis results
    ht="<tt><table border=1>"
    for f in collectPastAnalyses():
        ht=ht+"<tr><td><a href='reports/"+f+"' target='_blank' rel='noopener noreferrer'>"+f+"</a></td></tr>"
    ht=ht+"</tt></table>"
    return ht


# In[ ]:


def livePage(deviceString): # generate a live preview of the data analysis with the current PNG figures
    thisSampleID=getThisRunSampleID(deviceString) # if there is a run that produces data, the run ID will exist
    thisSampleRef=readReferenceDefinition(thisSampleID).replace(".xlsx", "")
    cnvPlotPath="reports/"+thisSampleID+"_CNVplot.png"
    umapAllPlotPath="reports/"+thisSampleID+"_"+thisSampleRef+"_UMAP_all.png"
    umapAllPlotlyPath="reports/"+thisSampleID+"_"+thisSampleRef+"_UMAP_all.html"
    umapTopPlotPath="reports/"+thisSampleID+"_"+thisSampleRef+"_UMAP_top.png"
    ht="<html><body><tt>sample ID: "+thisSampleID+" with reference "+thisSampleRef+"</tt><br>"
    ht=ht+"<a href='"+cnvPlotPath+"' target='_blank'><img align='Top' src='"+cnvPlotPath+"' width='50%' alt='CNV plot will appear here'></a>"
    ht=ht+"<a href='"+umapAllPlotlyPath+"' target='_blank'><img align='Top' src='"+umapAllPlotPath+"' width='50%' alt='UMAP plot will appear here'></a>"
    ht=ht+"</tt></table><body></html>"
    return ht


# In[ ]:


def methcallLivePage(sampleName): # generate a self-refreshing page to invoke methylation calling
    ht="<html><head><title>MethCaller: "+sampleName+"</title>"
    ht=ht+"<meta http-equiv='refresh' content='3'></head><body>"
    ht=ht+"last refresh and console output at "+datetimestringnow()+"<hr>shell output<br><br><tt>"
    #ht=ht+calculateMethylationAndBamFromFast5Fastq(sampleName)
    ht=ht+f5cOneFast5(sampleName,analyzeOne=True)
    ht=ht+"</tt></body></html>"
    return ht


# In[ ]:


def menuheader(n,r): # generate a universal website header for the UI pages that contains a simple main menu
    menuList=[["index","Overview","General system information"],
              ["listPositions","Mk1b Status","Live status of all connected Mk1b devices"],
              ["startSequencing","Start seq.","Start a sequencing run on an idle Mk1b device"],
              ["startTestrun","Start test run","Start a test seq. run on an idle Mk1b device to verify that the previous flow cell wash was successful."],
              ["listExperiments","Seq. runs","List all buffered runs. Will be purged upon MinKNOW backend restart."],
              ["listRuns","Results","List all completed analysis results"],
              ["analyze","Analyze","Launch data analyses manually, e.g. for retrospective analysis"],
              ["about","About NanoDiP","Version, etc."]
             ]
    mm="<html><head><title>NanoDiP Version "+versionString+"</title>"
    if r>0: # autorefresh wanted
        mm=mm+"<meta http-equiv='refresh' content='"+str(r)+"'>"
    mm=mm+'''</head><body><table border=0 cellpadding=2><tr>
    <td><img src='img/EpiDiP_Logo_01.png' width='40px' height='40px'></td>'''
    nn=0
    for m in menuList:
        if n==nn:
            selectedColor="#E0E0E0"
        else:
            selectedColor="white"
        mm=mm+"<td bgcolor='"+selectedColor+"'><b><a href='"+m[0]+"' title='"+m[2]+"'>"+m[1]+"</a></b></td>"
        nn=nn+1
    mm=mm+"</tr></table><br>"
    return(mm)


# ### 6. CherryPy Web UI
# The browser-based user interface is based on CherryPy, which contains an intergrated web server and serves pages locally. Communication between the service and browser typically generates static web pages that may or may not contain automatic self refresh commands. In the case of self-refreshing pages, the browser will re-request a given page with leads to re-execution of the respective python functions. The main handles to these function are located in the Web UI cell below.

# In[ ]:


# Launch CherryPy Webserver. Relaunch this cell only unless other cells have been modified to restart.
class MinKnowIfPApi(object): # the CherryPy Web UI class that defines entrypoints and function calls
    # global variables within the CherryPy Web UI
    globalRunFolders=getRuns()
    globalRunStatus= [None] * len(globalRunFolders)
    cpgQueue=0
    umapQueue=0
    cnvpQueue=0
    
    @cherrypy.expose
    def index(self):
        myString=menuheader(0,15)
        myString=myString+"<tt><b>Computer:</b> "+str(socket.gethostname())+"</tt><br><br>"
        myString=myString+systemStats()
        return myString

    @cherrypy.expose
    def restartServer(self):
        myString=menuheader(0,15)
        myString=myString+"NanoDiP server is restarting. It may be necesscary to reload some tabs to resume operation. Click on a menu item to proceed."
        restartNanoDiP()
        return myString
    
    @cherrypy.expose
    def resetQueue(self,queueName=""):
        myString=menuheader(0,15)
        if queueName:
            if queueName=="cpg":
                MinKnowIfPApi.cpgQueue=0
            if queueName=="umap":
                MinKnowIfPApi.umapQueue=0
            if queueName=="cnvp":
                MinKnowIfPApi.cnvpQueue=0
            myString=myString+queueName+" queue reset"
        return myString
    
    @cherrypy.expose
    def listPositions(self):      
        myString=menuheader(1,10)
        positions=listMinionPositions()
        for pos in positions:
            n=str(pos.name) # pos.state does not tell much other than that the device is connected with USB ("running")
            myString=myString+"<br><iframe src='DeviceStatusLive?deviceString="+n+"' height='200' width='600' title='"+n+"' border=3></iframe>"
            myString=myString+"<iframe src='AnalysisStatusLive?deviceString="+n+"' height='200' width='600' title='"+n+"' border=3></iframe>"
            myString=myString+"<br><a href='DeviceStatusLive?deviceString="+n+"' target='_blank' title='Click to open device status page in new tab or window'>"+n+"</a>"
            myString=myString+", live state: "+getRealDeviceActivity(n)
            activeRun=getActiveRun(n)
            myString=myString+", active run: "+getActiveRun(n)
            if activeRun!="none":
                myString=myString+" <a href='launchAutoTerminator?sampleName="+getThisRunSampleID(n)+"&deviceString="+n+"' target='_blank'>"
                myString=myString+"<br>Click this link to launch automatic run terminator after "+str(round(wantedBases/1e6))+" MB.</a>"
                myString=myString+"<br><font color=''#ff0000'><a href='stopSequencing?deviceId="+n+"' title='Clicking this will terminate the current run immediately! Use with care!'>terminate manually</a></font>"
            myString=myString+"<br><br>"
        myString=myString+"</body></html>"
        return myString

    @cherrypy.expose
    def startSequencing(self,deviceId="",sampleId="",runDuration="",referenceFile=""):
        myString=menuheader(2,0)
        if sampleId:
            if float(runDuration)>=0.1:
                sys.argv = ['',
                            '--host','localhost',
                            '--position',deviceId,
                            '--sample-id',sampleId,
                            '--experiment-group',sampleId,
                            '--experiment-duration',runDuration,
                            '--basecalling',
                            '--fastq',
                            '--fastq-reads-per-file',readsPerFile,
                            '--fast5',
                            '--fast5-reads-per-file',readsPerFile,
                            '--verbose',
                            '--kit','SQK-RBK004',
                            '--barcoding',
                            '--barcode-kits','SQK-RBK004']
                realRunId=startRun()
                writeReferenceDefinition(sampleId,referenceFile)
                myString=myString+"sequencing run started for "+sampleId+" on "+deviceId+" as "+realRunId+" with reference "+referenceFile
                myString=myString+"<hr>"+getThisRunInformation(deviceId)
                myString=myString+"<hr><a href='launchAutoTerminator?sampleName="+sampleId+"&deviceString="+deviceId+"'>"
                myString=myString+"Click this link to launch automatic run terminator after "+str(round(wantedBases/1e6))+" MB.</a> "
                myString=myString+"If you do not start the run terminator, you will have to terminate the run manually, or it will stop after the predefined time."
        else:    
            myString=myString+'''<form action="startSequencing" method="GET">
                Select an idle Mk1b:&nbsp;<select name="deviceId" id="deviceId">'''
            positions=listMinionPositions()
            for pos in positions:
                thisPos=pos.name
                if getRealDeviceActivity(thisPos)=="idle":
                    if getFlowCellID(thisPos)!="":
                        myString=myString+'<option value="'+thisPos+'">'+thisPos+': '+getFlowCellID(thisPos)+'</option>'
            myString=myString+'''
                </select>&nbsp; and enter the sample ID:&nbsp;<input type="text" name="sampleId" />
                &nbsp;for&nbsp;<input type="text" name="runDuration" value="72" />&nbsp;hours.
                &nbsp;Reference set&nbsp;<select name="referenceFile" id="referenceFile">'''
            for ref in getReferenceAnnotations():
                myString=myString+'<option value="'+ref+'">'+ref+'</option>'
            myString=myString+'&nbsp;<input type="submit" value="start sequencing now"/></form>'
        return myString

    @cherrypy.expose
    def startTestrun(self,deviceId=""):
        myString=menuheader(3,0)
        if deviceId:
            sampleId=datetimestringnow()+"_TestRun_"+getFlowCellID(deviceId)
            sys.argv = ['',
                        '--host','localhost',
                        '--position',deviceId,
                        '--sample-id',sampleId,
                        '--experiment-group',sampleId,
                        '--experiment-duration','0.1',
                        '--basecalling',
                        '--fastq',
                        '--fastq-reads-per-file',readsPerFile,
                        '--fast5',
                        '--fast5-reads-per-file',readsPerFile,
                        '--verbose',
                        '--kit','SQK-RBK004',
                        '--barcoding',
                        '--barcode-kits','SQK-RBK004']
            realRunId=startRun()
            myString=myString+"sequencing run started for "+sampleId+" on "+deviceId+" as "+realRunId
            myString=myString+"<hr>"+getThisRunInformation(deviceId)
        else:    
            myString=myString+'''<form action="startTestrun" method="GET">
                Select an idle Mk1b:&nbsp;<select name="deviceId" id="deviceId">'''
            positions=listMinionPositions()
            for pos in positions:
                thisPos=pos.name
                if getRealDeviceActivity(thisPos)=="idle":
                    if getFlowCellID(thisPos)!="":
                        myString=myString+'<option value="'+thisPos+'">'+thisPos+': '+getFlowCellID(thisPos)+'</option>'
            myString=myString+'''
                </select>&nbsp;<input type="submit" value="start test run now (0.1h)"/></form>'''
        return myString
    
    @cherrypy.expose
    def stopSequencing(self,deviceId=""):      
        myString=menuheader(1,0)
        myString=myString + stopRun(deviceId)
        myString=myString + "<br><br>Click on any menu item to proceed."
        return myString
    
    @cherrypy.expose
    def listExperiments(self):
        myString=menuheader(4,10)
        myString=myString+"Running and buffered experiments:<br>"
        experiments=listMinionExperiments()
        myString=myString+experiments
        return myString 
    
    @cherrypy.expose
    def listRuns(self):
        myString=menuheader(5,0)
        myString=myString+makePastAnalysesTable()
        return myString    
      
    @cherrypy.expose
    def analyze(self):
        myString=menuheader(6,0)        
        myString=myString+analysisLaunchTable()
        return myString
 
    @cherrypy.expose
    def cnvplot(self, sampleName=None):
        myString=""
        if sampleName:
            while MinKnowIfPApi.cnvpQueue>0:
                time.sleep(2)
            MinKnowIfPApi.cnvpQueue+=1
            try:
                createCNVPlot(sampleName)
                errorString=""
            except:
                errorString="<b><font color='#FF0000'>ERROR OCCURRED, PLEASE RELOAD TAB</font></b>"
            myString="<html><head><title>"+sampleName+" at "+datetimestringnow()+"</title></head><body>"
            myString=myString+errorString
            myString=myString+"<img src='reports/"+sampleName+"_CNVplot.png' width='100%'>"
            myString=myString+"</body></html>"
            MinKnowIfPApi.cnvpQueue-=1
        return myString
    
    @cherrypy.expose
    def umapplot(self, sampleName=None, refAnno=None):
        myString=""
        if sampleName and refAnno:
            while MinKnowIfPApi.umapQueue>0:
                time.sleep(2)
            MinKnowIfPApi.umapQueue+=1
            refAnnoName=refAnno.replace(".xlsx","")
            try:
                methylationUMAP(sampleName,refAnno)
                errorString=""
            except:
                errorString="<b><font color='#FF0000'>ERROR OCCURRED, PLEASE RELOAD TAB</font></b>"
            myString="<html><head><title>"+sampleName+" against "+refAnno+" at "+datetimestringnow()+"</title>"
            myString=myString+"<meta http-equiv='refresh' content='1; URL=reports/"+sampleName+"_"+refAnnoName+"_UMAP_all.html'>"
            myString=myString+"</head><body>"
            myString=myString+errorString
            myString=myString+"Loading UMAP plot. It it fails, <a href='reports/"+sampleName+"_"+refAnnoName+"_UMAP_all.html'>click here to load plot</a>."
            myString=myString+"</body></html>"
            MinKnowIfPApi.umapQueue-=1
        return myString
    
    @cherrypy.expose
    def makePdf(self, sampleName=None, refAnno=None):    
        myString=""
        if sampleName and refAnno:
            myString=generatePdfReport(sampleName,refAnno)
        return myString
    
    @cherrypy.expose
    def about(self):
        myString=menuheader(7,0)+'''
        <b>NanoDiP</b> is a tool to obtain and analyze low-coverage whole genome
        nanopore sequencing information through bascalling, genomic alignment,
        copy number extrapolation, and unsupervised machine learning by UMAP-based
        dimensions reduction. It is the nanopore-centered implementation of
        <a href="http://www.epidip.org">EpiDiP</a> which stands for <i>Epigenomic
        Digital Pathology</i>. NanoDiP hence abbreviates <i>Nanopore
        Digital Pathology</i>.<br><br>
        Nanopore sequencing is developed and sold by <a href="https://nanoporetech.com/">ONT</a>.
        Then authors of this software are not affiliated with ONT.<br><br>
        This software is licensed under the 
        <a href="https://www.gnu.org/licenses/gpl-3.0.html">GPLv3</a>. By using this
        program, you agree to the terms specified herein.<br><br>
        <b>This software is not a medical device.</b> Its use occurs in the sole responsibility
        of the treating physician. The authors shall not be held liable for any
        damage caused by this software. 
        <b>Basic understanding of how this system works and internal validation are 
        strongly advised before implementation in a diagnostic setting.</b>
        '''
        return myString  
    
    @cherrypy.expose
    def DeviceStatusLive(self,deviceString=""):
        currentFlowCellId=getFlowCellID(deviceString)
        myString="<html><head><title>"+deviceString+": "+currentFlowCellId+"</title>"
        try:
            myString=myString+"<meta http-equiv='refresh' content='2'>"
            if getRealDeviceActivity(deviceString)=="sequencing":
                myString=myString+"<body bgcolor='#00FF00'>"
            else:
                myString=myString+"<body>"
            myString=myString+"<b>"+deviceString+": "+currentFlowCellId+"</b><br><tt>"
            myString=myString+getMinKnowApiStatus(deviceString)
        except:
            myString=myString+"<br>No previous device activity, information will appear as soon as the device has been running once in this session.<br>"
        myString=myString+"Sample ID: "+getThisRunSampleID(deviceString)+"<br>"
        myString=myString+getThisRunState(deviceString)
        myString=myString+"<br>"+getThisRunYield(deviceString)
        myString=myString+"</tt></body></html>"
        return myString
    
    @cherrypy.expose
    def AnalysisStatusLive(self,deviceString=""):
        myString=""
        if deviceString:
            myString=livePage(deviceString)
        return myString

    @cherrypy.expose
    def analysisLauncher(self,functionName="",sampleName="",refAnno=""):
        if functionName and sampleName and refAnno:
            myString="<html><head><title>"+sampleName+" "+functionName+"</title></head><body>"
            myString=myString+functionName+" launched for "+sampleName+" "
            if refAnno!="None":
                myString=myString+"against "+refAnno
            myString=myString+" at "+datetimestringnow()+". "
            myString=myString+"Frame below will display result upon completion, if this tab/window is kept open."
            if refAnno=="None":
                myString=myString+"<br><iframe src='./"+functionName+"?sampleName="+sampleName+"' height='95%' width='100%' title='"+sampleName+"' border=3></iframe>"
            else:
                myString=myString+"<br><iframe src='./"+functionName+"?sampleName="+sampleName+"&refAnno="+refAnno+"' height='95%' width='100%' title='"+sampleName+"' border=3></iframe>"
        else:
            myString="Nothing to launch. You may close this tab now."
        return myString
    
    @cherrypy.expose
    def analysisPoller(self,sampleName="",deviceString="",runId=""):
        myString="<html><head>"
        if sampleName and deviceString and runId:
                myString=myString+"<title>Poller: "+sampleName+"/"+deviceString+"/"+runId+"</title>"
                myString=myString+"<meta http-equiv='refresh' content='15'>"
                myString=myString+"<body>"
                myString=myString+"Last refresh for "+sampleName+"/"+deviceString+"/"+runId+" at "+datetimestringnow()
                myString=myString+"</body></html>"
                writeRunTmpFile(sampleName,deviceString)
        return myString
    
    @cherrypy.expose
    def methylationPoller(self,sampleName=""):
        while MinKnowIfPApi.cpgQueue>0:
            time.sleep(2)
        MinKnowIfPApi.cpgQueue+=1
        myString=methcallLivePage(sampleName)
        MinKnowIfPApi.cpgQueue-=1
        return myString

    @cherrypy.expose
    def launchAutoTerminator(self,sampleName="",deviceString=""):
        myString="ERROR"
        if sampleName and deviceString:
            myString=thisRunWatcherTerminator(deviceString,sampleName)
        return myString
    
# CherryPy server configuration
config = {  
    'global' : {
        'server.socket_host' : cherrypyHost,
        'server.socket_port' : cherrypyPort,
        'server.thread_pool' : cherrypyThreads,
        'response.timeout' : 10,
        'server.shutdown_timeout': 1 
              }
}
# Start CherryPy Webserver
if debugLogging==True:
    cherrypy.log.screen=True #set access logging
    cherrypy.config.update({'log.screen': True})
else:
    cherrypy.log.screen=False #set access logging
    cherrypy.config.update({'log.screen': False})
    cherrypy.config.update({ "environment": "embedded" })
print("NanoDiP server running at http://"+cherrypyHost+":"+str(cherrypyPort))
if __name__ == '__main__':
#    cherrypy.tree.mount(root, '/', config=config) 
    cherrypy.quickstart(MinKnowIfPApi(),
                        '/',
                        {'/favicon.ico':
                         {'tools.staticfile.on':True,
                          'tools.staticfile.filename':
                          thisFaviconPath},
                        '/img':
                         {'tools.staticdir.on': True,
                          'tools.staticdir.dir': 
                          imgPath},
                        '/reports':
                         {'tools.staticdir.on': True,
                          'tools.staticdir.dir': 
                          nanodipReportDir}})


# ### ^^^ LIVE LOG ABOVE ^^^ 
# All CherryPy access will be logged here, including live progress bars for computationally intense analyses. Detailed access logging is turned off by default (accessLogging is False), but can be turned on,e.g., for debugging, in the configuration section at the beginning of this notebook. While it is not required to have at look at these during normal operation, information contained in the log may be helpful in troubleshooting. Line numbers in error messages indicated here typically match those given in the respective Jupyter Notebook cells.
# 
# To preseve these messages, halt the Python kernel, save and close the notebook to send it for support. This makes sure that the code as well as the error messages will be preserved.
# 
# To launch the user interface, wait until you see a pink log entry that the web server has started, then navigate to http://localhost:8080.
